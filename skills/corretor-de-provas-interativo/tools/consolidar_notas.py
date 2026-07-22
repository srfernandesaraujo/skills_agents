#!/usr/bin/env python3
# coding: utf-8
"""
Consolida notas ja julgadas pelo assistente (uma prova por vez, no chat) em uma
planilha .xlsx com tres abas: Notas, Analise de Itens e Resumo da Turma.

Este script NUNCA julga ou reavalia uma resposta de aluno — ele apenas soma,
organiza e calcula estatisticas descritivas em cima de pontuacoes que ja foram
decididas questao por questao nas Etapas 1-5 da skill "corretor-de-provas-interativo".

Uso:
    python consolidar_notas.py <entrada.json> <saida.xlsx>

Formato exato do JSON de entrada:
{
  "titulo": "Prova de Biologia - 2o bimestre",          // opcional
  "nota_maxima": 10.0,                                    // opcional; default = soma dos valores das questoes
  "criterios_aprovacao": {"aprovado_pct": 60, "recuperacao_pct": 40},  // opcional, em % da nota_maxima
  "questoes": [
    {"nome": "Q1", "valor": 2.0},
    {"nome": "Q2 (dissertativa)", "valor": 3.0}
  ],
  "alunos": [
    {"nome": "Joao Silva", "versao": "A", "pontos": [2.0, 1.5]},   // "versao" e opcional
    {"nome": "Maria Souza", "pontos": [1.0, 3.0]}
  ]
}

"pontos" de cada aluno deve ter o mesmo tamanho e ordem de "questoes".
"""

import sys
import json
import argparse
import statistics as st

try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    from openpyxl.chart import BarChart, Reference
except ImportError:
    print("Este script requer openpyxl. Instale com: pip install -r requirements.txt", file=sys.stderr)
    raise

FILL_APROVADO = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FILL_RECUPERACAO = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_REPROVADO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_HEADER = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
FONT_HEADER = Font(color="FFFFFF", bold=True)
FILL_PROBLEMATICA = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_REVISAR = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")


def carregar_entrada(caminho):
    with open(caminho, "r", encoding="utf-8") as f:
        dados = json.load(f)

    questoes = dados.get("questoes", [])
    alunos = dados.get("alunos", [])
    if not questoes:
        raise ValueError("JSON de entrada precisa de ao menos uma questao em 'questoes'.")
    if not alunos:
        raise ValueError("JSON de entrada precisa de ao menos um aluno em 'alunos'.")

    n_questoes = len(questoes)
    for aluno in alunos:
        if len(aluno.get("pontos", [])) != n_questoes:
            raise ValueError(
                f"Aluno '{aluno.get('nome', '?')}' tem {len(aluno.get('pontos', []))} pontuacoes, "
                f"mas ha {n_questoes} questoes. As listas precisam ter o mesmo tamanho e ordem."
            )
    return dados


def status_aprovacao(pct, criterios):
    aprovado_pct = criterios.get("aprovado_pct", 60)
    recuperacao_pct = criterios.get("recuperacao_pct", 40)
    if pct >= aprovado_pct:
        return "Aprovado", FILL_APROVADO
    if pct >= recuperacao_pct:
        return "Recuperacao", FILL_RECUPERACAO
    return "Reprovado", FILL_REPROVADO


def classificar_dificuldade(pct_acerto):
    if pct_acerto > 75:
        return "Facil"
    if pct_acerto >= 40:
        return "Medio"
    return "Dificil"


def classificar_discriminacao(indice):
    if indice is None:
        return "N insuficiente (min. 4 alunos)", None
    if indice >= 0.40:
        return "Excelente", None
    if indice >= 0.30:
        return "Boa", None
    if indice >= 0.20:
        return "Regular - considerar revisao", FILL_REVISAR
    if indice >= 0:
        return "Fraca - revisar item", FILL_REVISAR
    return "Problematica - possivel erro no gabarito ou enunciado ambiguo", FILL_PROBLEMATICA


def calcular_discriminacao(alunos, idx_questao, valor_questao, totais):
    """Indice classico de discriminacao (Ebel): compara o desempenho medio dos
    27% melhores vs. 27% piores alunos (pelo total da prova) naquela questao especifica."""
    n = len(alunos)
    if n < 4:
        return None

    ordenados = sorted(range(n), key=lambda i: totais[i], reverse=True)
    grupo_tamanho = max(1, round(n * 0.27))
    grupo_superior = ordenados[:grupo_tamanho]
    grupo_inferior = ordenados[-grupo_tamanho:]

    media_superior = sum(alunos[i]["pontos"][idx_questao] for i in grupo_superior) / (grupo_tamanho * valor_questao)
    media_inferior = sum(alunos[i]["pontos"][idx_questao] for i in grupo_inferior) / (grupo_tamanho * valor_questao)
    return round(media_superior - media_inferior, 2)


def montar_aba_notas(wb, dados, nota_maxima, criterios):
    questoes = dados["questoes"]
    alunos = dados["alunos"]
    tem_versao = any("versao" in a for a in alunos)

    ws = wb.active
    ws.title = "Notas"

    header = ["Aluno"]
    if tem_versao:
        header.append("Versao")
    header += [q["nome"] for q in questoes] + ["Total", "%", "Status"]

    ws.append(header)
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    n_questoes = len(questoes)
    col_primeira_questao = 3 if tem_versao else 2
    col_total = col_primeira_questao + n_questoes
    col_pct = col_total + 1
    col_status = col_pct + 1

    for row_idx, aluno in enumerate(alunos, start=2):
        row = [aluno["nome"]]
        if tem_versao:
            row.append(aluno.get("versao", ""))
        row += list(aluno["pontos"])
        ws.append(row)

        col_ini = get_column_letter(col_primeira_questao)
        col_fim = get_column_letter(col_primeira_questao + n_questoes - 1)
        total_cell = ws.cell(row=row_idx, column=col_total)
        total_cell.value = f"=SUM({col_ini}{row_idx}:{col_fim}{row_idx})"

        pct_cell = ws.cell(row=row_idx, column=col_pct)
        pct_cell.value = f"={get_column_letter(col_total)}{row_idx}/{nota_maxima}*100"
        pct_cell.number_format = "0.0"

        total_calculado = sum(aluno["pontos"])
        pct_calculado = (total_calculado / nota_maxima) * 100 if nota_maxima else 0
        status_txt, fill = status_aprovacao(pct_calculado, criterios)
        status_cell = ws.cell(row=row_idx, column=col_status)
        status_cell.value = status_txt
        if fill:
            status_cell.fill = fill

    for col_cells in ws.columns:
        largura = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(largura, 40)

    ws.freeze_panes = "A2"
    return tem_versao, col_primeira_questao, col_total


def montar_aba_itens(wb, dados, nota_maxima):
    questoes = dados["questoes"]
    alunos = dados["alunos"]
    totais = [sum(a["pontos"]) for a in alunos]

    ws = wb.create_sheet("Analise de Itens")
    header = ["Questao", "Valor", "Media de acerto", "% acerto medio", "Dificuldade",
              "Indice de discriminacao", "Classificacao (discriminacao)"]
    ws.append(header)
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER
        cell.alignment = Alignment(horizontal="center")

    for idx, questao in enumerate(questoes):
        valor = questao["valor"]
        pontos_questao = [a["pontos"][idx] for a in alunos]
        media = st.mean(pontos_questao)
        pct_acerto = (media / valor) * 100 if valor else 0
        dificuldade = classificar_dificuldade(pct_acerto)
        indice = calcular_discriminacao(alunos, idx, valor, totais)
        classificacao, fill = classificar_discriminacao(indice)

        row_idx = ws.max_row + 1
        ws.append([
            questao["nome"], valor, round(media, 2), round(pct_acerto, 1),
            dificuldade, indice if indice is not None else "N/A", classificacao
        ])
        if fill:
            ws.cell(row=row_idx, column=7).fill = fill

    for col_cells in ws.columns:
        largura = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(largura, 45)
    ws.freeze_panes = "A2"


def montar_aba_resumo(wb, dados, nota_maxima):
    alunos = dados["alunos"]
    totais = [sum(a["pontos"]) for a in alunos]

    ws = wb.create_sheet("Resumo da Turma")
    ws.append(["Metrica", "Valor"])
    for cell in ws[1]:
        cell.fill = FILL_HEADER
        cell.font = FONT_HEADER

    media = st.mean(totais)
    mediana = st.median(totais)
    desvio = st.pstdev(totais) if len(totais) > 1 else 0
    ws.append(["Numero de alunos", len(totais)])
    ws.append(["Nota maxima possivel", nota_maxima])
    ws.append(["Media da turma", round(media, 2)])
    ws.append(["Mediana", round(mediana, 2)])
    ws.append(["Desvio padrao", round(desvio, 2)])
    ws.append(["Nota minima", round(min(totais), 2)])
    ws.append(["Nota maxima obtida", round(max(totais), 2)])

    ws.append([])
    ws.append(["Faixa de %", "Quantidade de alunos"])
    faixas = [(0, 20), (20, 40), (40, 60), (60, 80), (80, 100.001)]
    linha_inicio_faixas = ws.max_row + 1
    for ini, fim in faixas:
        qtd = sum(1 for t in totais if nota_maxima and ini <= (t / nota_maxima * 100) < fim)
        rotulo = f"{ini:.0f}-{fim if fim <= 100 else 100:.0f}%"
        ws.append([rotulo, qtd])
    linha_fim_faixas = ws.max_row

    if len(totais) >= 2:
        chart = BarChart()
        chart.title = "Distribuicao de notas (%)"
        chart.x_axis.title = "Faixa"
        chart.y_axis.title = "N de alunos"
        dados_ref = Reference(ws, min_col=2, min_row=linha_inicio_faixas, max_row=linha_fim_faixas)
        categorias_ref = Reference(ws, min_col=1, min_row=linha_inicio_faixas, max_row=linha_fim_faixas)
        chart.add_data(dados_ref, titles_from_data=False)
        chart.set_categories(categorias_ref)
        chart.legend = None
        ws.add_chart(chart, "E2")

    for col_cells in ws.columns:
        largura = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells) + 2
        ws.column_dimensions[col_cells[0].column_letter].width = min(largura, 40)


def gerar_planilha(caminho_entrada, caminho_saida):
    dados = carregar_entrada(caminho_entrada)
    questoes = dados["questoes"]
    criterios = dados.get("criterios_aprovacao", {})
    nota_maxima = dados.get("nota_maxima") or sum(q["valor"] for q in questoes)

    wb = Workbook()
    montar_aba_notas(wb, dados, nota_maxima, criterios)
    montar_aba_itens(wb, dados, nota_maxima)
    montar_aba_resumo(wb, dados, nota_maxima)
    wb.save(caminho_saida)
    return caminho_saida


def main():
    parser = argparse.ArgumentParser(description="Consolida notas ja julgadas em uma planilha .xlsx com analise de itens.")
    parser.add_argument("entrada", help="Caminho do JSON de entrada (ver docstring do script para o formato exato).")
    parser.add_argument("saida", help="Caminho do .xlsx de saida.")
    args = parser.parse_args()

    caminho = gerar_planilha(args.entrada, args.saida)
    print(json.dumps({"sucesso": True, "arquivo_gerado": caminho}, ensure_ascii=False))


if __name__ == "__main__":
    main()
